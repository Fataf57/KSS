from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.db.models import Sum, Count
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import Sale, SaleItem
from .serializers import SaleSerializer, SaleCreateSerializer, SaleListSerializer, SaleItemSerializer


class SaleViewSet(viewsets.ModelViewSet):
    queryset = Sale.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return SaleCreateSerializer
        elif self.action == 'list':
            return SaleListSerializer
        return SaleSerializer

    def get_queryset(self):
        queryset = Sale.objects.all()
        customer = self.request.query_params.get('customer', None)
        date_from = self.request.query_params.get('date_from', None)
        date_to = self.request.query_params.get('date_to', None)

        if customer:
            queryset = queryset.filter(customer_id=customer)
        if date_from:
            queryset = queryset.filter(sale_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(sale_date__lte=date_to)

        return queryset

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'])
    def export_report(self, request):
        """Génère un rapport Excel des ventes pour une période donnée"""
        date_from = request.query_params.get('date_from', None)
        date_to = request.query_params.get('date_to', None)

        queryset = Sale.objects.all()
        if date_from:
            queryset = queryset.filter(sale_date__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(sale_date__date__lte=date_to)

        queryset = queryset.order_by('sale_date')

        # Créer un workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport des Ventes"

        # En-têtes
        headers = ['Date', 'Heure', 'ID Vente', 'Produit', 'Quantité', 'Prix Unitaire', 'Total', 'Méthode de Paiement']
        ws.append(headers)

        # Style des en-têtes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Données
        total_general = 0
        for sale in queryset:
            sale_date = sale.sale_date.date()
            sale_time = sale.sale_date.time()
            for item in sale.items.all():
                row = [
                    sale_date.strftime('%d/%m/%Y'),
                    sale_time.strftime('%H:%M'),
                    sale.id,
                    item.product.name,
                    item.quantity,
                    float(item.unit_price),
                    float(item.subtotal),
                    sale.get_payment_method_display()
                ]
                ws.append(row)
                total_general += float(item.subtotal)

        # Ajouter le total
        ws.append([])
        ws.append(['TOTAL GÉNÉRAL', '', '', '', '', '', total_general, ''])

        # Style du total
        total_row = ws[ws.max_row]
        for cell in total_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="center")

        # Ajuster la largeur des colonnes
        column_widths = [12, 10, 10, 30, 10, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Enregistrer dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Créer la réponse HTTP
        filename = f"rapport_ventes_{date_from or 'all'}_{date_to or 'all'}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

